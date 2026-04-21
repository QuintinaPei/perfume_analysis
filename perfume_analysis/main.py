# coding: utf-8

# The Correlation Between Perfume Ingredients and Popularity
# Based on Data Mining and Text Analysis of the Top 10 Perfumes in 2024

# 1. Data Preprocessing
# First, we will split the consumer reviews for each perfume collected from the Fragrantica website into individual reviews and save them in Excel files. Since the collected text contains the word 'share' between every two reviews, we will use 'share' to split the reviews. Then, we will combine the reviews of ten perfumes into one dataset. This dataset will include the following information: perfume name, consumer review, and perfume ingredients.

from docx import Document
from openpyxl import Workbook
import pandas as pd
import re
import nltk
from nltk.corpus import stopwords
from nltk.stem import WordNetLemmatizer
from gensim import corpora
from gensim.models import LdaModel
from gensim.utils import simple_preprocess
from gensim.models.coherencemodel import CoherenceModel
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import LabelEncoder
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.svm import SVC
from sklearn.tree import DecisionTreeClassifier
from sklearn.metrics import accuracy_score, confusion_matrix, classification_report, roc_curve, auc
from tensorflow.keras.preprocessing.text import Tokenizer
from tensorflow.keras.preprocessing.sequence import pad_sequences
from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import Embedding, LSTM, Dense, SimpleRNN
from tensorflow.keras.utils import to_categorical
import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np
from wordcloud import WordCloud

def read_comments_from_docx(docx_file):
    doc = Document(open(docx_file, 'rb'))
    comments = []
    current_comment = []

    for paragraph in doc.paragraphs:
        if paragraph.text.strip() == "share":
            if current_comment:
                comments.append("\n".join(current_comment))
                current_comment = []
        else:
            current_comment.append(paragraph.text)

    if current_comment:
        comments.append("\n".join(current_comment))

    return comments

def write_comments_to_excel(comments, excel_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Comments"

    for idx, comment in enumerate(comments, start=1):
        ws.cell(row=idx, column=1, value=comment)

    wb.save(excel_file)

def merge_excel_files(excel_files, output_file):
    first_df = pd.read_excel(excel_files[0])
    columns = first_df.columns.tolist()

    # Merge data from all Excel files
    all_data = []
    for file in excel_files:
        df = pd.read_excel(file)
        all_data.append(df)

    # Combine the data into a DataFrame
    merged_df = pd.concat(all_data, ignore_index=True)

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        merged_df.to_excel(writer, index=False, header=False)  

        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        for idx, col in enumerate(columns):
            worksheet.cell(row=1, column=idx + 1, value=col)

def clean_text(text):
    # Remove numbers
    text = re.sub(r'\d+', '', text)
    # Remove verbs, personal pronouns, and adverbs
    text = " ".join(word for word in text.split() if word.lower() not in ["i", "you", "he", "she", "it", "we", "they", "am", "is", "are", "was", "were", "be", "been", "being", "have", "has", "had", "do", "does", "did", "shall", "will", "should", "would", "can", "could", "may", "might", "must", "ought", "about", "above", "across", "after", "against", "along", "among", "around", "at", "before", "behind", "below", "beneath", "beside", "between", "beyond", "but", "by", "despite", "down", "during", "except", "for", "from", "in", "inside", "into", "like", "near", "of", "off", "on", "onto", "out", "outside", "over", "past", "since", "through", "throughout", "till", "to", "toward", "under", "underneath", "until", "up", "upon", "with", "within", "without", "share", "really", "smell", "smells", "just", "ago", "hours", "note", "don't"])
    return text

def preprocess_text(text):
    # Check if the input is a string, if not, return an empty string
    if isinstance(text, float):
        return ''
    
    # Convert the text to lowercase
    text = text.lower()
    # Remove all "//:" and the words before them
    text = re.sub(r'\S*\/\/\S*', '', text)
    # Remove non-English characters
    text = re.sub(r'[^a-zA-Z\s]', '', text)
    # Remove punctuation
    text = re.sub(r'[^\w\s]', '', text)
    # Remove whitespace characters
    text = text.strip()
    # Remove stopwords and lemmatize, and validate if they are valid English words
    tokens = [lemmatizer.lemmatize(word) for word in text.split() if word not in stop_words and word in english_words]
    return ' '.join(tokens)

def compute_coherence_values(dictionary, doc_term_matrix, texts, limit, start=2, step=3):
    coherence_values = []
    model_list = []
    for num_topics in range(start, limit, step):
        model = LdaModel(doc_term_matrix, num_topics=num_topics, id2word=dictionary, passes=50)
        model_list.append(model)
        coherence_model = CoherenceModel(model=model, texts=texts, dictionary=dictionary, coherence='c_v')
        coherence_values.append(coherence_model.get_coherence())
    return model_list, coherence_values

if __name__ == "__main__":
    # Step 1: Extract comments from Word document
    print("Step 1: Extracting comments from Word document...")
    docx_file = "Top10.docx"  
    excel_file = "top10.xlsx"  
    comments = read_comments_from_docx(docx_file)
    write_comments_to_excel(comments, excel_file)
    print("Comments extracted and saved to Excel successfully!")
    
    # Step 2: Merge Excel files
    print("\nStep 2: Merging Excel files...")
    excel_files = ["top1.xlsx", "top2.xlsx", "top3.xlsx", "top4.xlsx", "top5.xlsx",
                   "top6.xlsx", "top7.xlsx", "top8.xlsx", "top9.xlsx", "top10.xlsx"]
    output_file = "merged_file.xlsx"
    merge_excel_files(excel_files, output_file)
    print("Excel files merged successfully!")
    
    # Step 3: Clean and preprocess the data
    print("\nStep 3: Cleaning and preprocessing data...")
    df = pd.read_excel('merged_file.xlsx')
    
    # Create DataFrame for all comments
    all_comments = pd.DataFrame(columns=['Perfume', 'Review', 'Ingredients'])
    for index, row in df.iterrows():
        perfume_name = row['Perfume']
        reviews = row['Review']
        ingredients = row['Ingredients']
        
        # Process reviews
        if isinstance(reviews, str):
            reviews = [review.strip() for review in reviews.split('\n') if review.strip()]
        else:
            reviews = []
        
        for review in reviews:
            all_comments = pd.concat([all_comments, pd.DataFrame({'Perfume': [perfume_name], 'Review': [review], 'Ingredients': [ingredients]})], ignore_index=True)
    
    # Save initial reviews
    output_excel_file = 'review.xlsx'
    all_comments.to_excel(output_excel_file, index=False)
    
    # Clean text
    df['Review'] = df['Review'].apply(clean_text)
    df.to_excel('cleaned_reviews.xlsx', index=False)
    
    # Download NLTK resources
    nltk.download('stopwords')
    nltk.download('wordnet')
    nltk.download('words')
    nltk.download('vader_lexicon')
    
    # Initialize NLTK tools
    lemmatizer = WordNetLemmatizer()
    stop_words = set(stopwords.words('english'))
    english_words = set(nltk.corpus.words.words())
    
    # Preprocess text
    df['Review'] = df['Review'].apply(preprocess_text)
    df.dropna(axis=0, how='any', inplace=True)
    df.to_excel('processed_reviews.xlsx', index=False)
    print("Data preprocessing completed!")
    
    # Step 4: Sentiment analysis
    print("\nStep 4: Performing sentiment analysis...")
    from nltk.sentiment.vader import SentimentIntensityAnalyzer
    sia = SentimentIntensityAnalyzer()
    
    comments = df['Review'].tolist()
    sentiments = []
    for comment in comments:
        scores = sia.polarity_scores(comment)
        if scores['compound'] >= 0.05:
            sentiment = 'positive'
        elif scores['compound'] <= -0.05:
            sentiment = 'negative'
        else:
            sentiment = 'neutral'
        sentiments.append(sentiment)
    
    df['sentiment'] = sentiments
    df.to_excel('review_with_sentiment.xlsx', index=False)
    print("Sentiment analysis completed!")
    
    # Step 5: Dataset description
    print("\nStep 5: Analyzing dataset...")
    df['Review_Length'] = df['Review'].apply(lambda x: len(x.split()))
    print("Dataset Info:")
    print(df.info())
    print("\nSummary Statistics:")
    print(df.describe())
    
    # Plot sentiment distribution
    plt.figure(figsize=(8, 6))
    sns.countplot(x='sentiment', data=df, palette='viridis')
    plt.title('Count of Sentiment')
    plt.xlabel('Sentiment')
    plt.ylabel('Count')
    plt.savefig('sentiment_distribution.png')
    plt.close()
    
    # Step 6: Build LDA topic model
    print("\nStep 6: Building LDA topic model...")
    processed_docs = [simple_preprocess(doc) for doc in df['Review']]
    dictionary = corpora.Dictionary(processed_docs)
    doc_term_matrix = [dictionary.doc2bow(doc) for doc in processed_docs]
    
    num_topics = 5
    lda_model = LdaModel(doc_term_matrix, num_topics=num_topics, id2word=dictionary, passes=50)
    for idx, topic in lda_model.print_topics(-1):
        print(f'Topic {idx}: {topic}')
    
    # Compute coherence score
    coherence_model_lda = CoherenceModel(model=lda_model, texts=processed_docs, dictionary=dictionary, coherence='c_v')
    coherence_lda = coherence_model_lda.get_coherence()
    print(f'Coherence Score: {coherence_lda}')
    
    # Step 7: Analyze perfume ingredients
    print("\nStep 7: Analyzing perfume ingredients...")
    
    # Positive sentiment ingredients
    positive_reviews = df[df['sentiment'] == 'positive']
    ingredient_freq = {}
    for index, row in positive_reviews.iterrows():
        ingredients = row['Ingredients'].split('\n')
        for ingredient in ingredients:
            ingredient_freq[ingredient] = ingredient_freq.get(ingredient, 0) + 1
    
    most_common_ingredient = max(ingredient_freq, key=ingredient_freq.get)
    sorted_ingredients = sorted(ingredient_freq, key=ingredient_freq.get, reverse=True)
    top_five_ingredients = sorted_ingredients[:5]
    print(f"Most common individual ingredient with positive sentiment: {most_common_ingredient}")
    print(f"Top five individual ingredients with positive sentiment: {top_five_ingredients}")
    
    # Step 8: Model training and evaluation
    print("\nStep 8: Training and evaluating models...")
    
    # Prepare data for model training
    label_encoder = LabelEncoder()
    df['Sentiment_Encoded'] = label_encoder.fit_transform(df['sentiment'])
    y = to_categorical(df['Sentiment_Encoded'])
    X_train, X_test, y_train, y_test = train_test_split(df['Review'], y, test_size=0.2, random_state=42)
    
    # LSTM Model
    print("\nTraining LSTM model...")
    tokenizer = Tokenizer()
    tokenizer.fit_on_texts(X_train)
    X_train_seq = tokenizer.texts_to_sequences(X_train)
    X_test_seq = tokenizer.texts_to_sequences(X_test)
    
    maxlen = 100
    X_train_pad = pad_sequences(X_train_seq, maxlen=maxlen, padding='post')
    X_test_pad = pad_sequences(X_test_seq, maxlen=maxlen, padding='post')
    
    lstm_model = Sequential()
    lstm_model.add(Embedding(len(tokenizer.word_index)+1, 128, input_length=maxlen))
    lstm_model.add(LSTM(128, dropout=0.2, recurrent_dropout=0.2))
    lstm_model.add(Dense(3, activation='softmax'))
    lstm_model.compile(loss='categorical_crossentropy', optimizer='adam', metrics=['accuracy'])
    lstm_history = lstm_model.fit(X_train_pad, y_train, batch_size=32, epochs=20, validation_data=(X_test_pad, y_test), verbose=0)
    lstm_loss, lstm_accuracy = lstm_model.evaluate(X_test_pad, y_test, verbose=0)
    print(f'LSTM Test Accuracy: {lstm_accuracy}')
    
    # SVM Model
    print("\nTraining SVM model...")
    vectorizer = TfidfVectorizer(max_features=5000)
    vectorizer.fit(X_train)
    train_features = vectorizer.transform(X_train)
    test_features = vectorizer.transform(X_test)
    
    y_train_labels = np.argmax(y_train, axis=1)
    y_test_labels = np.argmax(y_test, axis=1)
    
    svm_model = SVC(kernel='linear', random_state=42, probability=True)
    svm_model.fit(train_features, y_train_labels)
    svm_predicted_labels = svm_model.predict(test_features)
    svm_accuracy = accuracy_score(y_test_labels, svm_predicted_labels)
    print(f'SVM Accuracy: {svm_accuracy}')
    
    # Decision Tree Model
    print("\nTraining Decision Tree model...")
    tree_model = DecisionTreeClassifier(random_state=42)
    tree_model.fit(train_features, y_train_labels)
    tree_predicted_labels = tree_model.predict(test_features)
    tree_accuracy = accuracy_score(y_test_labels, tree_predicted_labels)
    print(f'Decision Tree Accuracy: {tree_accuracy}')
    
    print("\nAll models trained successfully!")